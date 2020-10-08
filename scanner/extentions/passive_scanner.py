import subprocess
import netifaces
import sys
import re
import openpyxl
import socket
import time
from time import gmtime, strftime
from IPy import IP
from database import readSqliteTable
from openpyxl.utils import get_column_letter
from general import get_mac_addr, format_multi_line
from networking.ethernet import Ethernet
from networking.ipv4 import IPv4
from networking.pcap import Pcap


def main():
    ethinterface()
    passive_scan()

def ethinterface():
        """
        Uses the python library netifaces to enumerate a list of the interfaces
        on the system.
        Presents a list of interfaces and prompt the use to select one.
        """
        iflist = netifaces.interfaces()
        print('Interfaces found')

        for index in range(len(iflist)):
            print (index, ':', iflist[index])

        interface = input('enter interface # ')
        interface = int(interface)
        interface = iflist[interface]
    #    interface = input('Enter an interface name if needed: ')
        print('interface selected is:', interface)
        print()
        return interface

def passive_scan():
    TAB_1 = '\t - '
    TAB_2 = '\t\t - '

    pcap = Pcap('capture.pcap')
    conn = socket.socket(socket.AF_PACKET, socket.SOCK_RAW, socket.ntohs(3))
    
    exception_ip = ['0.0.0.0', '127.0.0.1', ]
    exception_ttl = ['1']

    while True:
        try:
            raw_data, addr = conn.recvfrom(65535)
            pcap.write(raw_data)
            eth = Ethernet(raw_data)

            print('\nEthernet Frame:')
            print(TAB_1 + 'Destination: {}, Source: {}, Protocol: {}'.format(eth.dest_mac, eth.src_mac, eth.proto))
            ipv4 = IPv4(eth.data) 
            ip_address = IP(ipv4.src)
            ip_type = ip_address.iptype()
            if ip_type == 'PRIVATE':
                if ipv4.src not in exception_ip:
                    if str(ipv4.ttl) not in exception_ttl:
                        if eth.proto == 8:
                            print(TAB_1 + 'IPv4 Packet:')
                            print(TAB_2 + 'Version: {}, Header Length: {}, TTL: {},'.format(ipv4.version, ipv4.header_length, ipv4.ttl))
                            print(TAB_2 + 'Protocol: {}, Source: {}, Target: {}'.format(ipv4.proto, ipv4.src, ipv4.target))
                        
                            readSqliteTable(eth.src_mac, ipv4.src)
        except KeyboardInterrupt:
            readSqliteTable(eth.src_mac, ipv4.src)
            time.sleep(5)
            print('>>>>>>>>>>>>>>>>>>>>>>>.save file.<<<<<<<<<<<<<<<<<<<<<<<<<<<<,,,')
            break
    pcap.close()

def find_device(result):
    result_list = result[2:-4]
    for i in result_list:
            ip = re.search(r'^(.*?)\t', i).group()
            mac = re.search(r'\t(.*?)\t', i).group()
            vendor = re.search(r'([^\t]+)$', i).group()
            open_excel(mac,ip,vendor)
            print(ip,mac,vendor)
            

def open_excel(src_mac,src_ip,mac_vendor):
    
    find_mac = 0
    for row in range(1, sheet.max_row + 1):
        mac_cell = sheet.cell(row = row, column = 1)
        cell_value = mac_cell.value
        if cell_value == src_mac:
           find_mac = 1
           check_ip(row,sheet,src_ip)

    if find_mac == 0:
        new_mac = sheet.cell(row = sheet.max_row+1, column = 1)
        print(sheet.max_row+1)
        new_mac.value = src_mac
        check_ip(sheet.max_row,sheet,src_ip)
    
    
    select_file.save('output.xlsx')
#    print('>>>>>>>>>>>>>>>>>>>>>>>Save & Close File<<<<<<<<<<<<<<<<<<<<<<<<<<<<')

def check_ip(row,sheet,src_ip):
    for i in range(4 , 7):
        scan_time = time.ctime(time.time())
        ip_cell = sheet.cell(row = row, column = i)
        cell_value = ip_cell.value
        time_cell = sheet.cell(row = row, column = i+3)
        time_cell.value = scan_time
        if cell_value == src_ip:
            break

        elif cell_value is None:
            ip_cell.value = src_ip
            break


main()
